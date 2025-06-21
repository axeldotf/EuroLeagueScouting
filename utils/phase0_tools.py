import os
import re
import pandas as pd
from typing import Dict
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
import gdown


def extract_table(input_file: str, output_file: str, apply_column_mapping: bool = False) -> None:
    """
    Phase 0 – Extract and clean raw player stats from Excel.
    
    Parameters:
    - input_file: path to raw Excel file
    - output_file: path to save cleaned Excel
    - apply_column_mapping: whether to map messy column names to standard ones
    """
    column_mapping = {
        "2PT%": "2P%", "PF 100 Poss": "2PA 100 POSS", "2PTA": "2PA/G", "2PTM": "2PM/G",
        "3PT%": "3P%", "3PTA": "3PA/G", "3PTM": "3PM/G", "AST": "Ast/G", "BLK": "BLK/G",
        "DF 100 Poss": "DF 100 POSS", "DF": "DF/G", "DR": "DR/G", "FGA": "FGA/G",
        "FGM": "FGM/G", "FT 100 Poss": "FTA 100 POSS", "FTA": "FTA/G", "FTM": "FTM/G",
        "MIN": "Min/G", "IND NET RTG": "NET RTG", "OR": "OR/G", "PTS": "PTS/G",
        "TM NAME": "TEAM", "ST": "ST/G", "TO%": "TO Ratio", "TO RATIO": "TO Ratio", "TO": "TO/G", "TR": "TR/G",
        "VAL": "Val/G"
    }

    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    # Individua intestazione (cella "RNK")
    start_row, start_col = None, None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'RNK':
                start_row, start_col = cell.row, cell.column
                break
        if start_row:
            break
    if start_row is None:
        raise ValueError("[ERROR] 'RNK' header not found in input sheet.")

    original_header = [ws.cell(row=start_row, column=c).value for c in range(start_col, ws.max_column + 1)]
    header = [column_mapping.get(col, col) for col in original_header] if apply_column_mapping else original_header

    try:
        name_col_idx = header.index('NAME')
    except ValueError:
        raise ValueError("[ERROR] 'NAME' column not found in headers.")

    # Stima righe valide
    estimated_rows = 0
    for r in range(start_row + 1, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value is None for c in range(start_col, ws.max_column + 1)):
            break
        estimated_rows += 1

    table_data = [header]
    for r in tqdm(range(start_row + 1, start_row + 1 + estimated_rows), desc="Extracting rows"):
        row_values = [ws.cell(row=r, column=c).value for c in range(start_col, ws.max_column + 1)]
        if row_values[name_col_idx] is not None:
            table_data.append(row_values)

    # Salva file Excel pulito
    new_wb = Workbook()
    new_ws = new_wb.active
    for r_idx, row in enumerate(table_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            new_ws.cell(row=r_idx, column=c_idx).value = value or ""
    new_wb.save(output_file)

    # Salva anche come Pickle
    df = pd.DataFrame(table_data[1:], columns=table_data[0])
    df.to_pickle(output_file.replace('.xlsx', '.pkl'))


def process_table(input_dir: str, output_dir: str, season_to_download: str = None) -> None:
    """
    Processes selected seasonal Excel file + all-time stats.
    
    Parameters:
    - input_dir: root input directory (e.g., "src/raw")
    - output_dir: root output directory (e.g., "src/clean")
    - season_to_download: one of {"23_24", "24_25"}, or None
    """
    gdrive_ids = {
        "24_25": "1whZ-lDdmLHLBmiZXnYoBjDpk4_gpxQMh",
        "23_24": "1claPABNcnpU3yJRu51PjXE0etYxeATaR"
    }

    local_paths = {
        "24_25": os.path.join("src", "raw", "24 25", "Players stats.xlsx"),
        "23_24": os.path.join("src", "raw", "23 24", "Players stats.xlsx")
    }

    if season_to_download in gdrive_ids:
        file_id = gdrive_ids[season_to_download]
        file_path = local_paths[season_to_download]
        if not os.path.exists(file_path):
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            print(f"[INFO] Downloading {season_to_download.replace('_', '/')} stats file...")
            download_gdrive_file(file_id, file_path)
        else:
            print(f"[INFO] File already exists for {season_to_download}. Skipping download.")
    elif season_to_download:
        print(f"[WARNING] Unknown season '{season_to_download}'. Skipping download.")

    print("\n[INFO] Starting selected file processing...\n")

    files_to_process = []

    # Add selected season only
    if season_to_download == "23_24":
        files_to_process.append({
            'input_relative': os.path.join('23 24', 'Players stats.xlsx'),
            'output_filename': '23_24_Players_stats_clean.xlsx'
        })
    elif season_to_download == "24_25":
        files_to_process.append({
            'input_relative': os.path.join('24 25', 'Players stats.xlsx'),
            'output_filename': '24_25_Players_stats_clean.xlsx'
        })

    # Always process all-time files
    files_to_process.extend([
        {
            'input_relative': os.path.join('all time', 'Statistiche giocatori - All time adv.xlsx'),
            'output_filename': 'All_time_Players_adv_stats_clean.xlsx'
        },
        {
            'input_relative': os.path.join('all time', 'Statistiche giocatori - All time trad.xlsx'),
            'output_filename': 'All_time_Players_trad_stats_clean.xlsx'
        }
    ])

    for file in files_to_process:
        input_path = os.path.join(input_dir, file['input_relative'])
        output_path = os.path.join(output_dir, file['output_filename'])
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        apply_mapping = '23_24' in output_path or '24_25' in output_path
        print(f"[PROCESSING] {file['output_filename']}")
        print(f"  ├─ Source:      {input_path}")
        print(f"  ├─ Destination: {output_path}")
        print(f"  └─ Mapping:     {'Yes' if apply_mapping else 'No'}")

        try:
            extract_table(input_path, output_path, apply_column_mapping=apply_mapping)
            print(f"[SUCCESS] File processed.\n")
        except Exception as e:
            print(f"[ERROR] Failed: {e}\n")

    print("[DONE] Selected files processed.\n")

def merge_alltime(trad_path: str, adv_path: str, output_path: str) -> None:
    """
    Merges all-time traditional and advanced stats into a single dataset.
    """
    def read_file(path: str) -> pd.DataFrame:
        alt_path = path.replace('.xlsx', '.pkl')
        return pd.read_pickle(alt_path) if os.path.exists(alt_path) else pd.read_excel(path)

    trad = read_file(trad_path)
    adv = read_file(adv_path)

    for df in [trad, adv]:
        if 'NAME' in df.columns:
            df['NAME'] = df['NAME'].apply(lambda x: ' '.join(x.split(', ')[::-1]) if isinstance(x, str) and ', ' in x else x)

    merge_keys = ['RNK', 'NAME', 'TEAM', 'ROLE', 'NAT', 'AGE', 'GP', 'W/L', 'W%']
    merged_df = pd.merge(trad, adv, on=merge_keys, how='outer', suffixes=('', '_adv'))
    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]
    merged_df.drop(columns=[col for col in ['MIN', 'SEASON_adv', 'HEIGHT_adv'] if col in merged_df.columns], inplace=True)

    grouping_keys = ['ROLE', 'TEAM', 'NAT', 'AGE', 'GP', 'W/L', 'W%']
    consolidated = []
    for keys, group in tqdm(merged_df.groupby(grouping_keys, dropna=False), desc="Consolidating"):
        base = dict(zip(grouping_keys, keys))
        rest = group.drop(columns=grouping_keys)
        combined = rest.apply(lambda col: col.dropna().unique()[0] if col.nunique(dropna=True) == 1 else col.dropna().iloc[0])
        consolidated.append({**base, **combined.to_dict()})
    final_df = pd.DataFrame(consolidated)

    final_df.sort_values(by=['NAME', 'SEASON'], inplace=True, ignore_index=True)
    final_df['RNK'] = range(1, len(final_df) + 1)

    front_cols = final_df.columns[7:11].tolist()
    ordered = front_cols + [col for col in final_df.columns if col not in front_cols]
    final_df = final_df[ordered]

    final_df.to_excel(output_path, index=False)
    final_df.to_pickle(output_path.replace('.xlsx', '.pkl'))
    print(f"[DONE] Merged dataset saved: {output_path}")


def download_gdrive_file(file_id: str, output_path: str) -> None:
    """
    Downloads a file from Google Drive via gdown.
    """
    url = f"https://drive.google.com/uc?id={file_id}"
    gdown.download(url, output_path, quiet=False)
    print(f"[INFO] Download complete: {output_path}")
